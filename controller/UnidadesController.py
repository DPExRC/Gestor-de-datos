from datetime import time
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl import load_workbook
import pandas as pd
import tkinter as tk

from components.get_path_resources import get_path_resources
from components.show_messages import show_error, show_message, show_warning


class UnidadesController:
    def __init__(self, model, view, volver_a_ajustes_callback):
        self.model = model
        self.view = view
        self.volver_a_ajustes_callback = volver_a_ajustes_callback

        self.view.set_controller(self)
        self.modified_cells = set()
        self.current_file_path = None


        #headers1, all_data1 = self.model.predeterminado()
        headers, all_data = self.model.predeterminado()
        


        # Si los datos fueron obtenidos correctamente, actualizamos la tabla
        if headers and all_data:

            self.view.update_table(headers, all_data)


    def start_edit(self, event=None):
        """Iniciar la edición de una celda seleccionada."""
        item = self.view.tree.selection()[0]
        col = self.view.tree.identify_column(event.x)
        col_idx = int(col.replace("#", "")) - 1
        self.selected_row_idx = self.view.tree.index(item)
        self.selected_column = col_idx
        x, y, width, height = self.view.tree.bbox(item, column=col)
        value = self.view.tree.item(item)["values"][col_idx]
        
        self.current_entry = tk.Entry(self.view.tree)
        self.current_entry.insert(0, value)
        self.current_entry.place(x=x, y=y, width=width, height=height)
        self.current_entry.focus()
        self.current_entry.bind("<Return>", self.save_edit)
        self.current_entry.bind("<FocusOut>", self.cancel_edit)
        # Registrar la celda modificada
        self.modified_cells.add((self.selected_row_idx, self.selected_column))  # Usamos un set para evitar duplicados

    def new_encabezados(self, antes, despues):
        """Buscar en la primera fila (encabezados) y actualizar el nombre si lo encuentra; si no, añadirlo al final."""
        ruta_archivo = get_path_resources("Libro2.xlsx")
        wb = openpyxl.load_workbook(ruta_archivo)

        # Seleccionar la hoja activa
        sheet = wb.active
        try:
            
            antes = antes[0]
            # Cargar el archivo Excel

            # Verificar si 'antes' ya existe en la primera fila (encabezados)
            found = False
            for col_num, cell in enumerate(sheet[1], start=1):
                if cell.value == antes:
                    # Si el encabezado 'antes' es encontrado, actualizar su nombre
                    sheet.cell(row=1, column=col_num, value=despues)
                    found = True
                    break

            # Si no se encontró el encabezado 'antes', añadir el nuevo encabezado al final
            if not found:
                # Obtener la última columna con datos
                last_col = sheet.max_column  

                sheet.cell(row=1, column=last_col+1, value=despues)

            # Guardar el archivo con los cambios
            wb.save(ruta_archivo)

        except Exception as e:
            print(f"Hubo un problema al agregar el encabezado a Libro2.xlsx: {e}")


    def save_edit(self, event=None):
            """Guardar la edición de una celda y actualizar 'FECHA DIGITACION' si corresponde."""
            if not self.current_entry:
                return


            new_value = self.current_entry.get().strip()
            item = self.view.tree.selection()[0]
            col_idx = self.selected_column
            row_idx = self.selected_row_idx

            # Guardar el valor editado en el modelo de datos
            self.model.all_data[row_idx][col_idx] = new_value
            self.modified_cells.add((row_idx, col_idx))

            # Guardar el nuevo valor en la celda editada
            current_values = list(self.view.tree.item(item)["values"])
            if col_idx == 0:
                self.new_encabezados(current_values, new_value)

            current_values[col_idx] = new_value
            self.view.tree.item(item, values=current_values)


            row_index = self.view.tree.index(item)
            self.model.all_data[row_index] = current_values

            self.is_data_modified = True

            if self.current_entry:
                self.current_entry.destroy()
                self.current_entry = None


    def cancel_edit(self, event=None):
        """Cancelar la edición y cerrar el Entry."""
        if self.current_entry:
            self.current_entry.destroy()
            self.current_entry = None



        
    def save_to_file(self):
            """Guardar los datos modificados en el archivo Excel en la ruta actual."""
            if not self.current_file_path:
                show_error("Error", "No hay archivo seleccionado.")
                return

            try:
                # Verificar si hay datos modificados
                if not self.is_data_modified:
                    show_warning("Advertencia", "No hay cambios para guardar.")
                    return

                # Verificar si hay filas con el valor "default" en alguna de sus celdas
                invalid_rows = [row for row in self.model.all_data1 if any("default" in str(cell).lower() for cell in row)]
                if invalid_rows:
                    show_warning("Advertencia", "No se puede guardar el archivo debido a valores 'default' en los datos.")
                    return
                
                 # Agregar pausa para asegurar que el archivo se guarde completamente
                time.sleep(1) 

                # Exportar los datos al archivo base sobrescribiéndolo
                self.model.export_to_excel(self.model.all_data1, self.model.headers, self.current_file_path)

                show_message("Éxito", f"Archivo guardado en {self.current_file_path}.")

                # Restablecer la bandera de modificación
                self.is_data_modified = False

            except Exception as e:
                show_error("Error al guardar archivo", str(e))



    def export_to_excel(self):
            """Exportar los datos a un archivo Excel"""
            file_path = get_path_resources("Unidades.xlsx")

            if file_path:
                self.export(self.model.all_data, self.model.headers, file_path)
                show_message("Éxito", f"Datos exportados a {file_path}")

    def export(self, data, headers, file_path):
        """Exportar los datos modificados a un archivo Excel"""
        df = pd.DataFrame(self.model.all_data, columns=self.model.headers)  # Usar datos actualizados

        df.to_excel(file_path, index=False)

        wb = load_workbook(file_path)
        ws = wb.active

        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col if cell.value)
            column_letter = col[0].column_letter
            ws.column_dimensions[column_letter].width = max_length + 2

        wb.save(file_path)

    def actualizar(self):
        """Actualiza los datos y la vista de la tabla."""
        headers, all_data = self.model.obtener_datos()  # Obtener datos actualizados

        if headers and all_data:
            self.view.update_table(headers, all_data)  # Actualizar la vista
        else:
            self.view.show_warning("Advertencia", "No se encontraron datos para actualizar.")

    def add_row(self):
        """Añadir una nueva fila con valores predeterminados en la posición deseada, usando los índices originales y actualizando el Treeview."""
        if not self.model.headers:
            tk.messagebox.showerror("Error", "No se puede añadir una fila sin datos cargados.")
            return

        # Crear una nueva fila con valores 'default'
        new_row = ["default"] * len(self.model.headers)

        # Si no hay ninguna fila seleccionada, añadir al final
        self.model.all_data.append(new_row)
        new_index = len(self.model.all_data) + 1
        #self.model.original_indices.append(new_index)

        # Actualizar el Treeview con los datos completos (filtrados y originales)
        self.view.update_table(self.model.headers, self.model.all_data)


        self.is_data_modified = True

        tk.messagebox.showinfo("Éxito", "Nueva fila añadida correctamente.")

    def delete_row(self):
        # Obtener la fila seleccionada
        selected_item = self.view.tree.selection()  

        if selected_item:
            selected_values = list(self.view.tree.item(selected_item)["values"])

            # Normalizar valores: convertir NaN a '' en self.model.all_data
            normalized_data = [[x if not pd.isna(x) else '' for x in row] for row in self.model.all_data]

            if selected_values in normalized_data:

                # Eliminar la fila correspondiente
                index_to_remove = normalized_data.index(selected_values)
                del self.model.all_data[index_to_remove]

                # Eliminar de la tabla visual
                self.view.tree.delete(selected_item)

                print("Fila eliminada correctamente.")

            else:
                print("La lista no está en la lista de listas.")


