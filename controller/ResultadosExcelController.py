from collections import defaultdict
import os
import tkinter as tk
from tkinter import filedialog, messagebox
from datetime import datetime
import unicodedata
from unidecode import unidecode  # Importamos unidecode para eliminar tildes
from openpyxl.utils import get_column_letter, column_index_from_string




from openpyxl import Workbook, load_workbook
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
import openpyxl.utils as xlutils

import pandas as pd

from components.get_path_resources import get_path_resources
from components.show_messages import show_error, show_message, show_warning



class ResultadosExcelController:
    def __init__(self, model, view, volver_a_main_callback):
        self.model = model
        self.view = view
        self.volver_a_main_callback = volver_a_main_callback

        self.selected_file = None
        self.is_data_modified = False
        self.current_entry = None
        self.selected_row_idx = None
        self.selected_column = None

        self.view.set_controller(self)
        self.current_file_path = None
        self.modified_cells = set()
        self.selected_idx = set()
        self.idx1 = []


        # Intentar cargar el archivo predeterminado al abrir la vista
        self.cargar_archivo_predeterminado()

        # Asociar el evento de "KeyRelease" a los campos de filtro para activar la actualización automática
        self.view.bind_filter_event(self.filter_data)


    def cargar_archivo_predeterminado(self):
        """Carga el archivo predeterminado desde el archivo de directorios."""
        try:
            file_path = self.leer_directorio()  # Obtiene la ruta del archivo desde directorios.txt

            if not file_path or not os.path.exists(file_path):
                raise FileNotFoundError(f"No se encontró una ruta válida en directorios.txt: {file_path}")

            headers, data = self.model.load_file(file_path)
            self.view.update_table(headers, data)
            self.current_file_path = file_path

        except Exception as e:
            show_error("Error al cargar archivo", str(e))


    def select_file(self):
        """Abrir un cuadro de diálogo para seleccionar un archivo."""
        file_path = filedialog.askopenfilename(
            title="Seleccionar archivo",
            filetypes=[("Archivos Excel", ".xlsx"), ("Todos los archivos", ".*")]
        )

        try:
                headers, data = self.model.load_file(file_path)
                self.view.update_table(headers, data)
                self.current_file_path = file_path  # Actualiza la ruta del archivo actual
        except Exception as e:
                show_error("Error al cargar archivo", str(e))

    def load_file(self):
        """Muestra una tabla en Tkinter con solo los encabezados, sin cargar datos de un archivo."""
        headers = self.model.headers  # Obtener los encabezados definidos en el modelo
        self.view.update_table(headers, [])  # Llamar a la vista con una tabla vacía

    def start_edit(self, event=None):
        """Iniciar la edición de una celda seleccionada."""
        item = self.view.tree.selection()[0]
        col = self.view.tree.identify_column(event.x)
        col_idx = int(col.replace("#", "")) - 1
        self.selected_row_idx = self.view.tree.index(item)
        self.selected_column = col_idx

            # Lista de índices de columnas que no se pueden editar
        columnas_bloqueadas = [0, 1, 5]  # Ajusta según sea necesario (ejemplo: columna 0, 1 y 3 bloqueadas)

        if col_idx in columnas_bloqueadas:
            show_message("Información", f"Edición bloqueada para la columna")
            return  # No permite la edición

        x, y, width, height = self.view.tree.bbox(item, column=col)
        value = self.view.tree.item(item)["values"][col_idx]
        
        self.current_entry = tk.Entry(self.view.tree)
        self.current_entry.insert(0, value)
        self.current_entry.place(x=x, y=y, width=width, height=height)
        self.current_entry.focus()
        self.current_entry.bind("<Return>", self.save_edit)
        self.current_entry.bind("<FocusOut>", self.cancel_edit)

        # Registrar la celda modificada
        self.modified_cells.add((self.selected_row_idx, self.selected_column))  # Usamos un set para evitar duplicados

    def generar_archivo_mensual_controller(self):  

        self.model.loading_file()      
        # Llamar a funciones específicas del modelo
        headers, data = self.model.loading_file()
        
        if headers and data:
                df = pd.DataFrame(data, columns=headers)
                nombre_mes = datetime.now().strftime("%B").upper()
                año = datetime.now().strftime("%Y")
                default_filename = f"{nombre_mes}_{año}.xlsx"
                
                directorio = filedialog.askdirectory(title="Seleccionar ubicación para guardar")
                if directorio:
                    file_path = f"{directorio}/{default_filename}"
                    df.to_excel(file_path, index=False)

                    # Ajustar columnas y centrar datos
                    workbook = load_workbook(file_path)
                    sheet = workbook.active

                    # Ajustar ancho de columnas y centrar texto
                    for col in sheet.columns:
                        max_length = 10
                        column_letter = col[0].column_letter
                        for cell in col:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        sheet.column_dimensions[column_letter].width = max_length + 2

                    # Quitar formato de encabezados
                    for cell in sheet[1]:
                        cell.font = Font(bold=True)
                        cell.border = Border(
                            left=Side(style=None),
                            right=Side(style=None),
                            top=Side(style=None),
                            bottom=Side(style=None)
                        )

                    workbook.save(file_path)
        
                    show_message("Éxito", f"Archivo guardado en: {file_path}")
                    self.guardar_directorio(file_path)
                    return file_path
        else:
                show_error("Error", "No hay datos para generar el archivo.")
        
        return headers, data
    
    def guardar_directorio(self, file_path):

        """Escribe contenido en un archivo.
        Args:
            ruta_archivo: La ruta al archivo donde se va a escribir.
            contenido: El texto que se va a escribir en el archivo.
        """
        file_path_direct = get_path_resources("directorios.txt")

        try:
            with open(file_path_direct, 'w') as f:  # Abre el archivo en modo escritura ('w')
                f.write(file_path + "\n")  # Escribe el contenido
            print(f"Se ha escrito en el archivo")

        except Exception as e:
            print(f"Error al escribir en el archivo: {e}")
    

    def leer_directorio(self):
        """Lee la última línea de un archivo de texto y muestra todo el contenido en caso de error."""
        try:
            file_path_leer = get_path_resources("directorios.txt")
            
            
            if not os.path.exists(file_path_leer):
                return None

            with open(file_path_leer, 'r') as f:
                lineas = f.readlines()

            if lineas:
                return lineas[-1].strip()  # Última línea sin espacios
            else:
                # Leer todo el contenido y mostrarlo en el error
                with open(file_path_leer, 'r') as f:
                    contenido = f.read()
                messagebox.showerror("Error", f"Error. Contenido del archivo:\n{contenido if contenido else 'El archivo está vacío'}")
                return None

        except Exception as e:
            print(f"Error al leer el archivo de directorio: {e}")
            return None


    def vacios(self):
        """Muestra las filas donde las columnas 'FECHA RECEPCION', 'FECHA DIGITACION' y 'RESULTADO' están vacías o contienen NaN, según el estado del checkbox."""

        # Verificar si el checkbox está marcado
        if self.view.select_all_var_vacios.get() == 1:  # Si está activado
            vacias = []
            headers = self.model.headers

            # Asegurarse de que las columnas requeridas existan
            try:
                fecha_recepcion_idx = headers.index("FECHA RECEPCION")
                fecha_digitacion_idx = headers.index("FECHA DIGITACION")
                resultado_idx = headers.index("RESULTADO")
            except ValueError:
                show_error("Error", "Las columnas 'FECHA RECEPCION', 'FECHA DIGITACION' o 'RESULTADO' no se encuentran en los datos.")
                return

            # Recorrer las filas y verificar si están vacías o contienen NaN
            for idx, row in enumerate(self.model.all_data):
                # Verificar si las celdas de las columnas específicas están vacías o contienen NaN
                fecha_recepcion_vacia = pd.isna(row[fecha_recepcion_idx]) or (isinstance(row[fecha_recepcion_idx], str) and row[fecha_recepcion_idx].strip() == "")
                fecha_digitacion_vacia = pd.isna(row[fecha_digitacion_idx]) or (isinstance(row[fecha_digitacion_idx], str) and row[fecha_digitacion_idx].strip() == "")
                
                # Para 'RESULTADO', verificamos si el valor es None, NaN o está vacío
                resultado_vacio = pd.isna(row[resultado_idx]) or (isinstance(row[resultado_idx], str) and row[resultado_idx].strip() == "") or (isinstance(row[resultado_idx], float) and row[resultado_idx] != row[resultado_idx])  # NaN check

                # Si las tres columnas están vacías o contienen NaN, agregar la fila a la lista de vacías
                if fecha_recepcion_vacia and fecha_digitacion_vacia and resultado_vacio:
                    vacias.append(self.model.all_data[idx])

            # Si hay filas vacías, mostrarlas en la vista
            if vacias:
                self.view.update_table(self.model.headers, vacias)  # Actualiza la tabla con las filas vacías
                show_message("Filas vacías encontradas", f"Se han encontrado {len(vacias)} filas con fechas y resultados vacíos.")
            else:
                show_message("No hay filas vacías", "No se encontraron filas con fechas o resultados vacíos.")
        else:
            # Si el checkbox está desmarcado, eliminar las filas vacías de la vista (si es necesario)
            self.view.update_table(self.model.headers, self.model.all_data)  # Actualiza la tabla con todas las filas
            show_message("Todas las filas visibles", "Se muestran todas las filas.")

    
    def rangos(self):
        file = get_path_resources("Rangos.xlsx")
        try:
            df = pd.read_excel(file)
            return df
        except Exception as e:
            print(e)
            


    def save_edit(self, event=None):
        """Guardar la edición de una celda y actualizar 'FECHA DIGITACION' si corresponde."""
        if not self.current_entry:
            return


        headers = ["LOCALIDAD", "MUESTRA", "FECHA MUESTRA", "FECHA RECEPCION", "FECHA DIGITACION", "ANALISIS", "RESULTADO", "UNIDAD"]
        new_value = self.current_entry.get().strip()
        item = self.view.tree.selection()[0]
        col_idx = self.selected_column
        row_idx = self.selected_row_idx #fila seleccionada en el treeview
        col_name = self.model.headers[col_idx]
        
        data = self.model.df

        #print(f"sis: {data.index[1]}")
        fila1 = data.iloc[0].to_list()
        # Filtrar las filas que coinciden con current_values

        # Obtener la fila original de la tabla
        current_values = list(self.view.tree.item(item)["values"])

        
        selected_values = self.view.tree.item(item)["values"]
        for idx, row in enumerate(self.model.all_data):
            # Comparar la fila seleccionada con cada fila en all_data
            if all(
                (value == selected_value or (pd.isna(value) and pd.isna(selected_value)))
                for value, selected_value in zip(row, selected_values)
            ):
                ## indice
                self.selected_idx = idx
                self.idx1.append(idx)
        ##print("Índices únicos:", self.print_idx1())
        #print("Índices seleccionados:", self.idx1)  # Imprimir la lista completa



        # Asegurar que la cantidad de valores en la fila coincida con los headers
        if len(current_values) > len(self.model.headers):
            current_values = current_values[:len(self.model.headers)]

        # Actualizar el valor editado
        current_values[col_idx] = new_value
        self.view.tree.item(item, values=current_values)

        # Buscar índices de columnas
        if "RESULTADO" in self.model.headers and "FECHA DIGITACION" in self.model.headers:
            resultados_idx = self.model.headers.index("RESULTADO")
            fecha_digitacion_idx = self.model.headers.index("FECHA DIGITACION")

            # Si se edita "RESULTADO", actualizar "FECHA DIGITACION"
            if col_idx == resultados_idx:
                current_values = list(self.view.tree.item(item)["values"])
                resultado_value = current_values[resultados_idx]

                # Si "RESULTADO" está vacío, vaciar "FECHA DIGITACION", si no, actualizar con la fecha actual
                if not resultado_value:
                    current_values[fecha_digitacion_idx] = ""
                else:
                    fecha_actual = datetime.now().strftime("%d/%m/%Y")  # Fecha en formato corto
                    current_values[fecha_digitacion_idx] = fecha_actual

                # Actualizar el valor de la fila
                self.view.tree.item(item, values=current_values)
                self.is_data_modified = True



        # Validaciones para distintos tipos de datos
        if col_name in ["FECHA MUESTRA", "FECHA RECEPCION", "FECHA DIGITACION"]:
            if new_value:  # Si hay un valor, validar formato de fecha
                try:
                    formato = "%d/%m/%Y" if col_name == "FECHA MUESTRA" else "%d/%m/%Y"
                    fecha = pd.to_datetime(new_value, format=formato, errors="coerce")

                    if pd.isna(fecha):
                        show_warning("Advertencia", f"Formato incorrecto en {col_name} (debe ser {formato.replace('%H:%M', 'HH:MM')})")
                        return
                    if col_name != "FECHA MUESTRA":
                        new_value = fecha.strftime("%d/%m/%Y")
                        self.is_data_modified = True

                except Exception:
                    show_error("Error", f"Error al convertir {col_name}")
                    return
            else:
                respuesta = messagebox.askyesno(
                    "Confirmación", f"La celda en fila {row_idx + 1}, columna {col_name} está vacía.\n¿Desea guardarla así?"
                )
                if not respuesta:
                    return

        elif col_name == "RESULTADO":  # Validación de rangos
            if new_value:

                try:

                    valor_float = float(new_value)
                    if valor_float.is_integer():
                        new_value = int(valor_float)
                    else:
                        new_value = round(valor_float, 6)


                    # Verificación de rangos
                    localidad = self.view.tree.item(item)["values"][headers.index("LOCALIDAD")]
                    punto_muestreo = self.view.tree.item(item)["values"][headers.index("MUESTRA")]
                    analisis = self.view.tree.item(item)["values"][headers.index("ANALISIS")]

                    df_rangos = self.rangos()
                    fila_rango = df_rangos[(df_rangos["LOCALIDAD"] == localidad) &
                                        (df_rangos["MUESTRA"] == punto_muestreo) &
                                        (df_rangos["ANALISIS"] == analisis)]

                    if not fila_rango.empty:
                        minimo = fila_rango.iloc[0]["MINIMO"]
                        maximo = fila_rango.iloc[0]["MAXIMO"]

                        if pd.notna(minimo) and valor_float < minimo:
                            print(valor_float)
                            respuesta = messagebox.askyesno("Advertencia", f"El valor ingresado ({valor_float}) es menor al mínimo ({minimo}).\n¿Desea guardarlo?")
                            if not respuesta:
                                return

                        if pd.notna(maximo) and valor_float > maximo:
                            respuesta = messagebox.askyesno("Advertencia", f"El valor ingresado ({valor_float}) es mayor al máximo ({maximo}).\n¿Desea guardarlo?")
                            if not respuesta:
                                return
                        
                        self.is_data_modified = True

                except Exception as e:
                    show_error("Error inesperado", f"{type(e).__name__}: {e}")
                    return

                except ValueError:
                    if not isinstance(new_value, (int, float)):  
                                show_error("Formato incorrecto", f"Formato incorrecto en: {col_name}.\nModificar valor: {new_value} DE TIPO : {type(new_value)}")                    
                                return
    
        self.modified_cells.add((row_idx, col_idx))

        # Guardar el nuevo valor en la celda editada
        current_values = list(self.view.tree.item(item)["values"])
        ###data

        current_values[col_idx] = new_value
        self.view.tree.item(item, values=current_values)


        # Si no hay índice original, usar la posición en el treeview (puede ser incorrecto)
        row_index = self.view.tree.index(item)
        #print(row_index)
        #print("-------------------------------DATOS-------------------------------")

        self.model.all_data[self.selected_idx] = current_values
        #print(self.selected_idx)
        #print(self.model.all_data[idx])

        self.is_data_modified = True

        if self.current_entry:
            self.current_entry.destroy()
            self.current_entry = None



    def cancel_edit(self, event=None):
        """Cancelar la edición y cerrar el Entry."""
        if self.current_entry:
            self.current_entry.destroy()
            self.current_entry = None

    def export_to_excel(self):
        """Exportar los datos a un archivo Excel"""
        file_path = filedialog.asksaveasfilename(
            title="Guardar como",
            defaultextension=".xlsx",
            filetypes=[("Archivos Excel", "*.xlsx")]
        )

        if file_path:
            # Crear el archivo Excel
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Datos"

            # Agregar encabezados
            for col_num, header in enumerate(self.model.headers, start=1):
                cell = sheet.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style=None),
                    right=Side(style=None),
                    top=Side(style=None),
                    bottom=Side(style=None)
                )

            # Agregar datos
            for row_num, row_data in enumerate(self.model.all_data, start=2):
                for col_num, value in enumerate(row_data, start=1):
                    cell = sheet.cell(row=row_num, column=col_num, value=value)
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            # Ajustar ancho de columnas
            for col in sheet.columns:
                max_length = 10
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                sheet.column_dimensions[col_letter].width = max_length + 2

            # Guardar el archivo
            workbook.save(file_path)
            show_message("Éxito", f"Datos exportados a {file_path}")

    def save_to_file(self):
        """Guardar los datos modificados en el archivo Excel en la ruta actual."""
        if not self.current_file_path:
            show_error("Error", "No hay archivo seleccionado.")
            return

        try:
            
            datos_procesados, unique_indices = self.print_idx1()  # Recibir ambos valores
            print("Datos Procesados:", datos_procesados)  # Mostrar datos
            print("Índices Únicos:", unique_indices)  # Mostrar índices únicos
    
            # Verificar si hay datos modificados
            if not self.is_data_modified:
                show_warning("Advertencia", "No hay cambios para guardar.")
                return

            # Verificar si hay filas con el valor "default" o valores no deseados
            invalid_rows = [row for row in self.model.all_data if any("default" in str(cell).lower() for cell in row)]
            if invalid_rows:
                show_warning("Advertencia", "No se puede guardar el archivo debido a valores 'default' en los datos.")
                return

            # Crear un libro de trabajo y una hoja de trabajo con openpyxl
            wb = Workbook()
            sheet = wb.active
            sheet.title = "Datos"

            # Agregar encabezados
            for col_num, header in enumerate(self.model.headers, start=1):
                cell = sheet.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Agregar datos
            for row_num, row_data in enumerate(self.model.all_data, start=2):
                for col_num, value in enumerate(row_data, start=1):
                    cell = sheet.cell(row=row_num, column=col_num, value=value)
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            # Ajustar ancho de columnas
            for col in sheet.columns:
                max_length = 10
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                sheet.column_dimensions[col_letter].width = max_length + 2

            # Guardar el libro de trabajo en la ruta seleccionada
            wb.save(self.current_file_path)

            # Restablecer las celdas modificadas
            self.modified_cells.clear()

            show_message("Éxito", f"Archivo guardado en {self.current_file_path}.")

            # Restablecer la bandera de modificación
            self.is_data_modified = False

            # Resetear la lista después de guardar
            self.idx1.clear()

        except Exception as e:
            show_error("Error al guardar archivo", str(e))



    def filter_data(self, event=None):
        """Filtrar los datos según las entradas en los filtros, manteniendo la correspondencia con sus índices originales."""
        # Obtener todos los datos e índices originales del modelo
        all_data = self.model.all_data

        # Aplicar cada filtro sobre los datos
        for col_idx, filter_entry in enumerate(self.view.filters):
            search_term = filter_entry.get().lower().strip()  # Obtener el término de búsqueda
            if search_term:
                # Si hay un término para filtrar
                all_data = [row for row in all_data if search_term in str(row[col_idx]).lower()]

        # Actualizar la tabla con los datos filtrados
        self.view.update_table(self.model.headers, all_data)

        # Si el checkbox de "Seleccionar todas las filas visibles" está marcado, seleccionar las filas visibles
        if self.view.select_all_var.get():
            self.select_all_rows()

    def select_all_rows(self):
        """Seleccionar o deseleccionar todas las filas visibles según el estado del checkbox."""
        select_all = self.view.select_all_var.get()  # Obtener el estado del checkbox

        for item in self.view.tree.get_children():  # Iterar sobre todas las filas visibles
            self.view.tree.selection_add(item) if select_all else self.view.tree.selection_remove(item)



    def reset_filters(self):
        """Restablecer los filtros y mostrar todos los datos"""
        # Limpiar todos los filtros
        for filter_entry in self.view.filters:
            filter_entry.delete(0, tk.END)
        
        # Mostrar todos los datos
        self.view.update_table(self.model.headers, self.model.all_data)  # Mostrar todos los datos sin filtrar
   
    def add_row(self):
        """Añadir una nueva fila con valores predeterminados en la posición deseada, usando los índices originales y actualizando el Treeview."""
        if not self.model.headers:
            tk.messagebox.showerror("Error", "No se puede añadir una fila sin datos cargados.")
            return

        # Crear una nueva fila con valores 'default'
        new_row = ["default"] * len(self.model.headers)

        try:
            # Obtener la fila seleccionada en el Treeview (índice filtrado)
            selected_item = self.view.tree.selection()[0]
            row_index = self.view.tree.index(selected_item)  # Índice en la tabla filtrada

            # Si no hay filtro, insertar la fila después de la fila seleccionada
            if row_index >= len(self.model.all_data):
                tk.messagebox.showerror("Error", "No se encontró la fila en la tabla.")
                return
            
            self.model.all_data.insert(row_index + 1, new_row)
            new_index = self.model.original_indices[row_index] + 1  # Generar el nuevo índice después de la seleccionada
            self.model.original_indices.insert(row_index + 1, new_index)

            self.is_data_modified = True


        except IndexError:
            # Si no hay ninguna fila seleccionada, añadir al final
            self.model.all_data.append(new_row)
            new_index = max(self.model.all_data) + 1  # Generar el nuevo índice para el final
            self.model.original_indices.append(new_index)

        # Actualizar el Treeview con los datos completos (filtrados y originales)
        self.view.update_table(self.model.headers, self.model.all_data)

        # Si hay un filtro, actualizar los datos mostrados en el Treeview con la fila añadida
        if self.view.filters:
            self.filter_data()  # Filtrar los datos para reflejar la fila añadida en la vista filtrada

        tk.messagebox.showinfo("Éxito", "Nueva fila añadida correctamente.")


   
    def print_all_data(self):
        """Verifica si al menos un filtro tiene un valor y devuelve True si hay filtros activos, False si no."""
        return any(filter_entry.get().strip() for filter_entry in self.view.filters)


    def delete_row(self):
        """Eliminar las filas seleccionadas con confirmación y actualizar el archivo base."""
        try:
            # Obtener las filas seleccionadas en el Treeview
            selected_items = self.view.tree.selection()  # Devuelve una lista de los elementos seleccionados

            if not selected_items:
                messagebox.showerror("Error", "Por favor, selecciona al menos una fila para eliminar.")
                return

            # Confirmación de eliminación
            confirm = messagebox.askyesno("Confirmación", "¿Estás seguro de que deseas eliminar estas filas?")
            if not confirm:
                return

            # Ordenar las filas seleccionadas para asegurarse de eliminar desde el índice más alto
            selected_items = sorted(selected_items, key=lambda item: self.view.tree.index(item), reverse=True)

            # Iterar sobre las filas seleccionadas para eliminar
            for selected_item in selected_items:
                row_index = self.view.tree.index(selected_item)  # Índice en el TreeView (puede estar filtrado)
                
                # Obtener todas las filas del Treeview y convertirlas a DataFrame
                filas = [self.view.tree.item(item)["values"] for item in self.view.tree.get_children()]
                df_treeview = pd.DataFrame(filas, columns=self.model.df.columns)


                if self.is_data_modified is True:
                    selected_values = self.view.tree.item(selected_item)["values"]
                    for idx, row in enumerate(self.model.all_data):
                        # Comparar la fila seleccionada con cada fila en all_data
                        if all(
                            (value == selected_value or (pd.isna(value) and pd.isna(selected_value)))
                            for value, selected_value in zip(row, selected_values)
                        ):
                            print("")
                            ##print(f"Fila repetida encontrada en el índice {idx}:")
                            ##print(f"tree: {selected_values}\nall_data: {row}")

                # Buscar la fila seleccionada en el DataFrame
                resultado = self.model.df[(self.model.df == list(self.view.tree.item(selected_item)["values"])).all(axis=1)]

                if not resultado.empty:
                    idx = resultado.index[0]  # Obtener el índice de la primera coincidencia
                else:
                    messagebox.showerror("Error", "No se pudo encontrar la fila en los datos originales.")
                    return

                # Asegurarse de que el índice sea válido
                if idx < len(self.model.all_data):
                    row_dataframe = self.model.all_data[idx]  # Obtener la fila original

                    #print(f"Datos seleccionados para eliminar: {row_dataframe}")
                    # Eliminar la fila del modelo y del TreeView
                    del self.model.all_data[idx]
                    #del self.model.original_indices[idx]
                    self.view.tree.delete(selected_item)
                    # Intentar eliminar la fila en el archivo Excel
                    if self.delete_rows_in_file([idx]):
                        print(f"Fila {idx} eliminada correctamente del archivo base.")
                    else:
                        messagebox.showerror("Error", f"No se pudo eliminar la fila {idx} del archivo base.")
                        return  # Si falla la eliminación en el archivo Excel, no continuar

            messagebox.showinfo("Éxito", "Filas eliminadas correctamente.")
            self.is_data_modified = True  # Marcar que se realizaron modificaciones

        except IndexError:
            messagebox.showerror("Error", "Ocurrió un error al intentar eliminar las filas seleccionadas.")

    def delete_rows_in_file(self, rows_to_delete):
        """
        Elimina las filas especificadas del archivo Excel original.

        Args:
            rows_to_delete (list): Lista de índices de fila a eliminar (1-indexados).
                Se recomienda omitir la fila de encabezados (por ejemplo, usar índices >=2).

        Returns:
            bool: True si la operación fue exitosa, False en caso de error.
        """
        try:
            # Obtiene la ruta del archivo desde directorios.txt
            file_path = self.leer_directorio()  
            # Imprimir la ruta del archivo base que se utilizará para eliminar filas

            # Cargar el libro de Excel usando openpyxl
            wb = load_workbook(file_path)
            ws = wb.active

            # Ordenar los índices de filas a eliminar en orden descendente para evitar problemas de reindexación
            for row_idx in sorted(rows_to_delete, reverse=True):
                ws.delete_rows(row_idx)

            # Guardar el libro sobrescribiendo el archivo original
            wb.save(file_path)
            wb.close()
            return True
        except Exception as e:
            print(f"Error al eliminar filas en el archivo: {e}")
            return False
            
    def volver_a_main(self):
        """Método para volver a la vista principal."""
        self.volver_a_main_callback()

    def normalizar_claves(self):
    # Leer el archivo Excel

        ruta = get_path_resources("Libro2.xlsx")

        df = pd.read_excel(ruta, usecols=["MUESTRA"])

        # Normalizar las tildes y eliminar duplicados
        def normalizar_texto(texto):
            return ''.join(c for c in unicodedata.normalize('NFD', str(texto).upper()) if unicodedata.category(c) != 'Mn')

        muestras_unicas = set(df["MUESTRA"].dropna().apply(normalizar_texto))

        #print(muestras_unicas)  # Muestra el conjunto sin duplicados y sin tildes

        return muestras_unicas  # Devuelve los valores si es necesario


    def actualizar_excel(self):
        """Carga las localidades con sus archivos si existen y busca coincidencias exactas en la fila 2."""
        ruta_guardado = get_path_resources("DirectoriosLocalidades.txt")

        # Obtener palabras clave desde normalizar_claves
        palabras_clave = self.normalizar_claves()

        try:
            with open(ruta_guardado, "r", encoding="utf-8") as file:
                for linea in file:
                    partes = linea.strip().split(": ", 1)
                    if len(partes) == 2:
                        localidad, ruta = partes
                        if ruta != "Sin asignar" and os.path.exists(ruta) and ruta.endswith(".xlsx"):  
                            print(f"Leyendo archivo para Localidad: {localidad}, Ruta: {ruta}")
                            try:
                                wb = openpyxl.load_workbook(ruta)
                                sheet = wb.active  # Seleccionar la hoja activa

                                # Buscar coincidencias exactas en la fila 2
                                columnas_encontradas = []
                                for col in range(1, sheet.max_column + 1):  # Recorrer columnas
                                    valor = sheet.cell(row=2, column=col).value  # Obtener valor en fila 2
                                    if valor and str(valor).strip() in palabras_clave:  # Comparación exacta
                                        letra_columna = xlutils.get_column_letter(col)  # Obtener letra de columna
                                        columnas_encontradas.append((letra_columna, valor))  # Guardar letra de columna y nombre

                                # Mostrar resultados
                                if columnas_encontradas:
                                    print(f"Palabras clave encontradas en {ruta}:")
                                    for letra, valor in columnas_encontradas:
                                        print(f"Columna {letra}: {valor}")
                                else:
                                    print(f"No se encontraron coincidencias exactas en {ruta}.")

                            except Exception as e:
                                show_error(f"Error al leer {ruta}", str(e))
        except Exception as e:
            show_error("Error", str(e))




    def obtener_ruta_localidad(self, localidad):
        """Lee el archivo DirectoriosLocalidades.txt y devuelve la ruta de la localidad especificada."""
        ruta_txt = get_path_resources("DirectoriosLocalidades.txt")

        if not os.path.exists(ruta_txt):
            print("Error: No se encontró el archivo DirectoriosLocalidades.txt")
            return None

        with open(ruta_txt, "r", encoding="utf-8") as file:
            for line in file:
                partes = line.strip().split(":",1)  # Suponiendo que los datos están separados por ";"

                if len(partes) == 2 and partes[0].strip() == localidad:
                    return partes[1].strip()  # Retorna la ruta del archivo de la localidad
        
        print(f"Error: No se encontró la ruta para la localidad '{localidad}' en DirectoriosLocalidades.txt")
        return None

    def print_idx1(self):
        seen = set()
        unique_indices = []
        resultados = defaultdict(list)  # Diccionario para agrupar por localidad, muestra y análisis

        # Cargar Rangos.xlsx
        ruta_rangos = get_path_resources("Rangos.xlsx")
        if not os.path.exists(ruta_rangos):
            print("Error: No se encontró el archivo Rangos.xlsx")
            return
        
        df_rangos = pd.read_excel(ruta_rangos)

        # Convertir a diccionario para búsqueda rápida
        ubicaciones_dict = {(row["LOCALIDAD"].strip(), row["MUESTRA"].strip(), row["ANALISIS"].strip()): 
                            row["UBICACION"] for _, row in df_rangos.iterrows()}
        

        for idx in self.idx1:
            if idx not in seen:
                seen.add(idx)
                unique_indices.append(idx)


        for idx in unique_indices:
            dato = self.model.all_data[idx]
            localidad = dato[0].strip()
            muestra = dato[1].strip()
            analisis = dato[5].strip()
            resultado = dato[6]
            valor_busqueda = dato[2].strip()  # Se usa la primera columna para buscar la fila

            # Buscar la ubicación en Rangos.xlsx
            key = (localidad, muestra, analisis)
            ubicacion = ubicaciones_dict.get(key, None)

            if not ubicacion:
                print(f"Advertencia: No se encontró la ubicación para {localidad} - {muestra} - {analisis}")
                continue

            # Obtener la ruta del archivo de la localidad
            ruta_localidad = self.obtener_ruta_localidad(localidad)
            if not ruta_localidad or not os.path.exists(ruta_localidad):
                print(f"Error: No se encontró el archivo en la ruta {ruta_localidad}")
                continue

            try:
                # Cargar el archivo Excel de la localidad
                wb = load_workbook(ruta_localidad)
                ws = wb.active

                # Buscar la fila donde la primera columna coincide con valor_busqueda
                fila_coincidente = None
                for row in range(5, ws.max_row + 1):
                    celda_valor = ws.cell(row=row, column=1).value  # Primera columna (A)

                    # Convertir fechas a string si es necesario
                    if isinstance(celda_valor, datetime):
                        celda_valor = celda_valor.strftime("%d/%m/%Y")

                    print(f"Fila {row}: {celda_valor} (Tipo: {type(celda_valor)})")
                    if celda_valor and str(celda_valor).strip() == valor_busqueda:
                        fila_coincidente = row
                        break

                if not fila_coincidente:
                    print(f"Advertencia: No se encontró '{valor_busqueda}' en la primera columna de {ruta_localidad}")
                    continue

                # Convertir la columna 'ubicacion' a índice numérico
                col_ubicacion = column_index_from_string(ubicacion)

                # Insertar el resultado en la fila encontrada y la columna de ubicación
                ws.cell(row=fila_coincidente, column=col_ubicacion, value=resultado)

                # Guardar los cambios en el archivo
                wb.save(ruta_localidad)
                print(f"✔ Resultado '{resultado}' insertado en '{ubicacion}' (Fila {fila_coincidente}) para {localidad} - {muestra} - {analisis}")

            except Exception as e:
                print(f"Error al actualizar el archivo {ruta_localidad}: {e}")

        return resultados, unique_indices  # Retorna los datos agrupados



    def actualizar_excel2(self):
        """
        Lee las rutas de DirectoriosLocalidades.txt, busca dato[1] en la fila 2, y luego busca dato[5] en la fila siguiente dentro de la misma columna y columnas hacia la derecha.
        """
        ruta_guardado = get_path_resources("DirectoriosLocalidades.txt")
        resultados = {}

        try:
            with open(ruta_guardado, "r", encoding="utf-8") as file:
                for linea in file:
                    partes = linea.strip().split(": ", 1)
                    if len(partes) == 2:
                        localidad, ruta = partes
                        if ruta != "Sin asignar" and os.path.exists(ruta) and ruta.endswith(".xlsx"):  
                            print(f"Leyendo archivo para Localidad: {localidad}, Ruta: {ruta}")
                            try:
                                wb = openpyxl.load_workbook(ruta)
                                sheet = wb.active  # Seleccionar la hoja activa

                                # Obtener los resultados de print_idx1
                                resultados_idx1 = self.print_idx1()
                                print(resultados_idx1)

                                # Iterar sobre los datos ya procesados en print_idx1
                                for dato in resultados_idx1:
                                    palabra_clave = str(dato[0][1]).strip().upper()  # Convertir dato[1] a mayúsculas
                                    print(palabra_clave)
                                    referencia = str(dato[0][3]).strip()  # Valor a buscar en la fila siguiente (dato[5] en tu lista)
                                    print(referencia)
                                    col_encontrada = None

                                    # Paso 1: Buscar la palabra clave (dato[1]) en la fila 2
                                    for col in range(1, sheet.max_column + 1):
                                        celda = sheet.cell(row=2, column=col).value
                                        if celda and str(celda).strip().upper() == palabra_clave:
                                            col_encontrada = col
                                            break  # Se encontró la columna, salimos del bucle

                                    if col_encontrada:
                                        # Paso 2: Bajar una fila (fila 3) y buscar dato[5] en las columnas a la derecha de la columna encontrada
                                        encontrado = False
                                        for fila in range(3, sheet.max_row + 1):  # Comenzamos desde la fila 3
                                            for col in range(col_encontrada, sheet.max_column + 1):  # Buscar en la columna encontrada y hacia la derecha
                                                celda_actual = sheet.cell(row=fila, column=col).value
                                                if celda_actual and str(celda_actual).strip() == referencia:
                                                    # Guardamos la fila y columna en la que se encontró
                                                    resultados[localidad] = (fila, col)
                                                    encontrado = True
                                                    break
                                            if encontrado:
                                                break
                                        if not encontrado:
                                            resultados[localidad] = -1  # Si no encuentra dato[5], almacena -1
                                    else:
                                        resultados[localidad] = -1  # Si no encuentra dato[1], almacena -1

                            except Exception as e:
                                show_error(f"Error al leer {ruta}", str(e))
                                resultados[localidad] = -1
        except Exception as e:
            show_error("Error", str(e))

        return resultados  # Devuelve el diccionario con los resultados
