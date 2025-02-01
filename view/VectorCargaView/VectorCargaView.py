import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from tkinter import messagebox
from openpyxl import load_workbook
import pandas as pd
import threading
import os

from components.botones import BotonBaseGrid


class VectorCargaView:
    def __init__(self, root, main):
        self.root = root
        self.main = main
        self.vector_carga_view = tk.Frame(root)
        self.is_data_modified = False
        self.auto_save_interval = 10000  # 10 segundos
        self.auto_save_thread = None
        self.selected_file = None 



        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

        # Frame principal
        self.main_frame = tk.Frame(self.root)
        self.main_frame.pack(fill="both", expand=True, padx=10, pady=10)

        # Frame para los filtros
        self.filter_frame = tk.Frame(self.main_frame)
        self.filter_frame.pack(side="top", fill="x", pady=5)

        # Frame para los botones
        self.button_frame = tk.Frame(self.main_frame)
        self.button_frame.pack(side="top", fill="x", pady=5)

        # Botones
        BotonBaseGrid(
            self.button_frame,
            select_file=self.select_file,
            save_command=self.save_to_file,
            reset_command=self.reset_filters,
            add_row_command=self.add_row,
            delete_row_command=self.delete_row,
            export_command=self.export_to_excel,
            volver=self.volver

        )

        # Frame para la tabla
        self.table_frame = tk.Frame(self.main_frame)
        self.table_frame.pack(fill="both", expand=True)

        # Treeview para mostrar datos
        self.tree = ttk.Treeview(self.table_frame, show="headings")
        self.tree.pack(fill="both", expand=True)
        self.tree.bind("<Double-1>", self.start_edit)  # Doble clic para editar

        # Atributos para manejar datos
        self.selected_file = None
        self.headers = []
        self.all_data = []
        self.filters = []
        self.current_entry = None
        self.selected_row_idx = None
        self.selected_column = None

    def volver(self):
        """Función para volver a la vista principal."""
        self.limpiar_main_frame()
        
        self.vector_carga_view.destroy()  # Elimina los widgets actuales
        self.main()  # Llama a la función que muestra la vista principal

    def limpiar_main_frame(self):
        """Elimina todos los widgets del frame principal."""
        for widget in self.root.winfo_children():
            widget.destroy()

    def select_file(self):
        """Abrir un cuadro de diálogo para seleccionar un archivo .xlsx"""
        file_path = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[("Archivos de Excel", "*.xlsx")]
        )
        if file_path:
            self.selected_file = file_path
            self.load_file()

    def load_file(self):
        """Cargar datos del archivo Excel seleccionado y procesarlos"""
        if self.selected_file:
            # Leer el archivo Excel en un DataFrame
            df = pd.read_excel(self.selected_file)

            # Especificar las columnas que deben fusionarse
            analysis_columns = [
                "DQO", "ST", "SST", "SSV", "ph", "AGV (ácido acético)",
                "alcalinidad (CaCO3)", "% humedad", "transmitancia"
            ]

            # Crear la nueva columna "Análisis"
            if any(col in df.columns for col in analysis_columns):
                df["Análisis"] = df[analysis_columns] \
                    .apply(lambda row: ", ".join(row.dropna().astype(str)), axis=1)
                df = df.drop(columns=[col for col in analysis_columns if col in df.columns])  # Eliminar columnas originales

            # Reorganizar para que "Análisis" sea la última columna
            columns_order = [col for col in df.columns if col != "Análisis"] + ["Análisis"]
            df = df[columns_order]

            # Guardar los encabezados y datos
            self.headers = list(df.columns)
            self.all_data = df.values.tolist()

            # Actualizar la tabla
            self.update_table(self.all_data)

            return df

    def update_table(self, data):
        """Actualizar la tabla con datos proporcionados"""
        self.tree.delete(*self.tree.get_children())  # Limpiar la tabla

        # Configurar encabezados y filtros
        if not self.tree["columns"]:
            self.tree["columns"] = list(range(len(self.headers)))
            for idx, header in enumerate(self.headers):
                self.tree.heading(idx, text=header)

                # Crear un Frame para contener el Entry y el Label
                filter_frame = tk.Frame(self.filter_frame, relief="groove", bd=1)
                filter_frame.grid(row=0, column=idx, padx=5, pady=5, sticky="w")

                # Crear el Label para mostrar el nombre del encabezado
                label = tk.Label(filter_frame, text=f"{header}:", anchor="w")
                label.pack(side="top", fill="x", padx=5, pady=2)

                # Crear el Entry para el filtro
                filter_entry = tk.Entry(filter_frame)
                filter_entry.pack(side="top", fill="x", padx=5, pady=2)

                # Asociar el evento de filtrado
                filter_entry.bind("<KeyRelease>", self.filter_data)
                self.filters.append(filter_entry)

        # Insertar datos
        for row in data:
            self.tree.insert("", "end", values=row)


    def start_edit(self, event):
        """Iniciar la edición de una celda seleccionada."""
        if self.current_entry:
            self.cancel_edit()

        # Obtener la celda seleccionada
        item = self.tree.selection()[0]
        col = self.tree.identify_column(event.x)  # Columna seleccionada (ej: '#1')
        col_idx = int(col.replace("#", "")) - 1  # Convertir a índice de columna
        self.selected_row_idx = self.tree.index(item)  # Índice de la fila seleccionada
        self.selected_column = col_idx

        # Obtener coordenadas de la celda seleccionada
        x, y, width, height = self.tree.bbox(item, column=col)

        # Obtener el valor actual de la celda
        value = self.tree.item(item)["values"][col_idx]

        # Crear un Entry para editar
        self.current_entry = tk.Entry(self.tree)
        self.current_entry.insert(0, value)  # Colocar el valor actual en el Entry
        self.current_entry.place(x=x, y=y, width=width, height=height)
        self.current_entry.focus()
        self.current_entry.bind("<Return>", self.save_edit)
        self.current_entry.bind("<FocusOut>", self.cancel_edit)

    def save_edit(self, event=None):
        """Guardar el valor editado en la celda."""
        new_value = self.current_entry.get()  # Obtener el nuevo valor
        self.tree.set(self.tree.selection()[0], column=self.selected_column, value=new_value)  # Actualizar en el Treeview

        # Convertir fila a lista para poder modificarla
        row_data = list(self.all_data[self.selected_row_idx])
        row_data[self.selected_column] = new_value  # Actualizar el valor en la fila
        self.all_data[self.selected_row_idx] = row_data  # Guardar los cambios

        self.is_data_modified = True

        self.cancel_edit()

    def cancel_edit(self, event=None):
        """Cancelar la edición y cerrar el Entry."""
        if self.current_entry:
            self.current_entry.destroy()
            self.current_entry = None

    def filter_data(self, event=None):
        """Filtrar los datos según las entradas en los filtros"""
        filtered_data = self.all_data

        for col_idx, filter_entry in enumerate(self.filters):
            search_term = filter_entry.get().lower().strip()
            if search_term:
                filtered_data = [
                    row for row in filtered_data
                    if search_term in str(row[col_idx]).lower()
                ]

        # Actualizar la tabla con los datos filtrados
        self.update_table(filtered_data)

    def reset_filters(self):
        """Restablecer los filtros y mostrar todos los datos"""
        for filter_entry in self.filters:
            filter_entry.delete(0, tk.END)
        self.update_table(self.all_data)

    def save_to_file(self):
        """Guardar los datos modificados en el archivo Excel."""
        if not self.selected_file:
            tk.messagebox.showerror("Error", "No hay archivo seleccionado.")
            return

        # Verificar si alguna fila contiene el valor "default"
        invalid_rows = [row for row in self.all_data if "default" in row]

        if invalid_rows:
            # Mostrar un mensaje de advertencia si hay filas con valores 'default'
            tk.messagebox.showwarning(
                "Advertencia",
                "No se puede guardar el archivo. Hay filas con valores 'default'. Corrige los datos antes de guardar."
            )
            return

        # Crear un DataFrame sin filas con valores 'default'
        filtered_data = [row for row in self.all_data if "default" not in row]
        df = pd.DataFrame(filtered_data, columns=self.headers)

        # Guardar el archivo Excel
        df.to_excel(self.selected_file, index=False)
        tk.messagebox.showinfo("Información", "Archivo guardado correctamente.")

    def export_to_excel(self):
        """Exportar los datos a un archivo Excel con formato específico, validando filas con 'default'."""
        if not self.all_data or not self.headers:
            tk.messagebox.showerror("Error", "No hay datos para exportar.")
            return

        # Verificar si hay filas con valores "default"
        invalid_rows = [row for row in self.all_data if "default" in row]
        if invalid_rows:
            tk.messagebox.showwarning(
                "Advertencia",
                "No se puede exportar el archivo. Hay filas con valores 'default'. Corrige los datos antes de exportar."
            )
            return

        # Crear un DataFrame con los datos actuales (sin modificar el contenido)
        df = pd.DataFrame(self.all_data, columns=self.headers)

        # Procesar las fechas si existe alguna columna relacionada
        if "PROGRAMA" in df.columns:
            try:
                dias_semana = {
                    'lunes': 0, 'martes': 1, 'miercoles': 2, 'jueves': 3, 'viernes': 4, 'sabado': 5, 'domingo': 6
                }
                reference_date = pd.Timestamp.now().normalize()
                inicio_semana = reference_date - pd.Timedelta(days=reference_date.weekday())

                def procesar_fechas(programa):
                    if not pd.isna(programa):
                        fechas = []
                        dias = programa.replace(" ", "").split(",")
                        for dia in dias:
                            if "-" in dia:
                                dia_inicio, dia_fin = dia.split("-")
                                dia_inicio = dia_inicio.strip().lower()
                                dia_fin = dia_fin.strip().lower()

                                if dia_inicio in dias_semana and dia_fin in dias_semana:
                                    fecha_inicio = inicio_semana + pd.Timedelta(days=dias_semana[dia_inicio])
                                    fecha_fin = inicio_semana + pd.Timedelta(days=dias_semana[dia_fin])
                                    fechas.append(f"{fecha_inicio.strftime('%d-%m-%Y')} a {fecha_fin.strftime('%d-%m-%Y')}")
                            else:
                                dia = dia.strip().lower()
                                if dia in dias_semana:
                                    fecha = inicio_semana + pd.Timedelta(days=dias_semana[dia])
                                    fechas.append(fecha.strftime("%d-%m-%Y"))
                        return ", ".join(fechas)
                    return programa

                df["PROGRAMA"] = df["PROGRAMA"].apply(procesar_fechas)
            except Exception as e:
                tk.messagebox.showerror("Error", f"Error al procesar las fechas: {e}")
                return

        # Abrir un cuadro de diálogo para seleccionar el nombre y ubicación del archivo
        file_path = filedialog.asksaveasfilename(
            title="Guardar como",
            defaultextension=".xlsx",
            filetypes=[("Archivos Excel", "*.xlsx")]
        )

        if not file_path:
            return  # El usuario canceló

        try:
            # Guardar el DataFrame en un archivo Excel con openpyxl
            with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="Datos")

                # Obtener el libro y la hoja activa
                workbook = writer.book
                worksheet = writer.sheets["Datos"]

                # Configurar el ancho de las columnas
                for column_cells in worksheet.columns:
                    max_length = 0
                    column_letter = column_cells[0].column_letter  # Obtener la letra de la columna
                    for cell in column_cells:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except Exception:
                            pass
                    adjusted_width = max_length + 2  # Ajustar el ancho con un margen adicional
                    worksheet.column_dimensions[column_letter].width = adjusted_width

            tk.messagebox.showinfo("Éxito", f"Datos exportados exitosamente a: {file_path}")
        except Exception as e:
            tk.messagebox.showerror("Error", f"No se pudo exportar el archivo: {e}")


    def add_row(self):
        """Añadir una nueva fila con valores predeterminados en la posición deseada."""
        if not self.headers:
            tk.messagebox.showerror("Error", "No se puede añadir una fila sin datos cargados.")
            return

        # Crear una nueva fila con valores 'default'
        new_row = ["default"] * len(self.headers)

        try:
            # Obtener la fila seleccionada
            selected_item = self.tree.selection()[0]
            selected_index = self.tree.index(selected_item)  # Índice de la fila seleccionada

            # Insertar la nueva fila en los datos justo después de la fila seleccionada
            self.all_data.insert(selected_index + 1, new_row)
        except IndexError:
            # Si no hay ninguna fila seleccionada, añadir al final
            self.all_data.append(new_row)

        # Actualizar la tabla con los datos
        self.update_table(self.all_data)

        tk.messagebox.showinfo("Éxito", "Nueva fila añadida correctamente.")

    
    def delete_row(self):
        """Eliminar la fila seleccionada."""
        try:
            # Obtener la fila seleccionada
            selected_item = self.tree.selection()[0]
            row_index = self.tree.index(selected_item)  # Índice de la fila seleccionada

            # Eliminar la fila de los datos y del Treeview
            del self.all_data[row_index]
            self.tree.delete(selected_item)

            tk.messagebox.showinfo("Éxito", "Fila eliminada correctamente.")
        except IndexError:
            tk.messagebox.showerror("Error", "Por favor, selecciona una fila para eliminar.")

    def on_close(self):
        """Método para verificar si hay cambios y preguntar antes de cerrar."""
        if self.is_data_modified:
            answer = messagebox.askyesnocancel("Confirmar", "Tienes cambios sin guardar. ¿Quieres guardarlos antes de salir?")
            if answer is None:  # Cancelar
                return
            if answer:  # Guardar antes de salir
                self.save_to_file()
        self.root.destroy()


    analysis_columns = [
            "DQO", "ST", "SST", "SSV", "ph", "AGV (ácido acético)",
            "alcalinidad (CaCO3)", "% humedad", "transmitancia"
            ]


    def asignar_columnas(df, analisis_split, analysis_columns):
        # Crear un DataFrame para almacenar las nuevas columnas asignadas
        with open("output.txt", "a") as f:
            f.write(f"df1: {df}\n")
        
        for col in analisis_split.columns:
            for i in range(len(analisis_split)):
                valor = analisis_split[col][i]
                if valor in analysis_columns:
                    # Asignar el valor a la columna correspondiente en el DataFrame
                    df.at[i, valor] = valor
        
        return df

    def dividir_analisis(df, analysis_columns, asignar_columnas):
        analisis_split = None
        if "analisis" in df.columns:
            # Dividir la columna 'analisis' en varias columnas basadas en las comas
            analisis_split = df["analisis"].str.split(',', expand=True)

            # Eliminar la columna 'analisis' original del DataFrame
            df = df.drop(columns=["analisis"])

            # Concatenar las nuevas columnas divididas al DataFrame original
            df = pd.concat([df, analisis_split], axis=1)
            
            # Asignar columnas según el contenido de las celdas
            df = asignar_columnas(df, analisis_split, analysis_columns)
            
            with open("output.txt", "a") as f:
                f.write(f"df: {df}\n")
                f.write(f"analisis_split: {analisis_split}\n")

        return df, analisis_split



    def auto_save(self):
        """Guarda automáticamente los datos en el archivo original con el mismo formato, respetando las columnas."""
        if self.selected_file:  # Si hay un archivo seleccionado
            try:
                # Verificar si el archivo ya existe
                if os.path.exists(self.selected_file):
                    # Abrir el archivo original para conservar el formato
                    wb = load_workbook(self.selected_file)
                    sheet = wb.active  # Suponiendo que trabajas con la hoja activa

                    # Crear un DataFrame con los datos que quieres guardar
                    df = pd.DataFrame(self.all_data, columns=self.headers)

                    # Procesar la columna de análisis y dividirla en múltiples columnas

                    analisis_split = None
                    if "analisis" in df.columns:
                        # Dividir la columna 'analisis' en varias columnas basadas en las comas
                        analisis_split = df["analisis"].str.split(',', expand=True)

                        # Eliminar la columna 'analisis' original del DataFrame
                        df = df.drop(columns=["analisis"])

                        # Concatenar las nuevas columnas divididas al DataFrame original
                        df = pd.concat([df, analisis_split], axis=1)
                        
                

                    # Actualizar los datos en el archivo original sin perder el formato
                    for i, row in df.iterrows():
                        for j, value in enumerate(row[:4]):
                            # Actualizar solo los valores de las celdas
                            sheet.cell(row=i+2, column=j+1, value=value)  # +2 para saltar la cabecera

                    # Ajustar el ancho de las columnas en función de los datos
                    for col in sheet.columns:
                        max_length = 0
                        column = col[0].column_letter  # Obtener la letra de la columna
                        for cell in col:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(cell.value)
                            except:
                                pass
                        adjusted_width = (max_length + 2)  # Agregar un margen extra para que las celdas no queden pegadas
                        sheet.column_dimensions[column].width = adjusted_width

                    # Guardar el archivo sin cambiar el formato
                    wb.save(self.selected_file)
                    print(f"Guardado en el archivo original, manteniendo formato: {self.selected_file}")
                else:
                    # Si el archivo original no existe, crear uno temporal
                    temp_file = self.selected_file + ".temp"  # Crear archivo temporal
                    df = pd.DataFrame(self.all_data, columns=self.headers)
                    df.to_excel(temp_file, index=False)
                    print(f"Archivo original no encontrado. Guardado en archivo temporal: {temp_file}")

            except Exception as e:
                print(f"Error al guardar automáticamente: {e}")
        else:
            print("No hay archivo seleccionado, no hay nada que guardar.")
        
        # Reiniciar el auto guardado
        self.start_auto_save()

    def iniciar_guardado(self):
        """Verifica si hay un archivo seleccionado para autoguardar"""
        if self.selected_file:
            print(f"Archivo seleccionado {self.selected_file}")
            self.start_auto_save()  # Inicia el autoguardado
        else:
            print("No hay archivo seleccionado, no hay nada que guardar")

    def start_auto_save(self):
        """Inicia el hilo de autoguardado en intervalos regulares si hay un archivo seleccionado."""
        print(f"select_file: {self.selected_file}")  # Imprimir para ver su valor

        if self.selected_file:
            print("Iniciando autoguardado...")
            self.auto_save_thread = threading.Timer(self.auto_save_interval / 1000, self.auto_save)
            self.auto_save_thread.start()
        else:
            print("No hay archivo seleccionado. El autoguardado no se ha iniciado.")


    def load_temp_file(self):
        """Cargar datos desde el archivo temporal si existe."""
        if not self.selected_file:
            print("No se ha seleccionado ningún archivo.")
            return  # Salir si no hay archivo seleccionado

        temp_file = self.selected_file + ".temp"
        if os.path.exists(temp_file):
            try:
                df = pd.read_excel(temp_file)
                self.headers = df.columns.tolist()
                self.all_data = df.values.tolist()
                self.update_table(self.all_data)  # Actualiza la tabla con los datos cargados
                print(f"Datos restaurados desde {temp_file}")
            except Exception as e:
                print(f"Error al cargar los datos: {e}")

    def stop_auto_save(self):
        """Detiene el hilo de autoguardado si está en ejecución."""
        if self.auto_save_thread and self.auto_save_thread.is_alive():
            print("Deteniendo el autoguardado.")
            self.auto_save_thread.cancel()
            self.auto_save_thread = None  # Reinicia el hilo

    def cerrar_vista(self):
        print("Cerrando la vista...")

        # Si no hay archivo seleccionado, simplemente se detiene el autoguardado
        if not self.select_file:
            print("No se seleccionó ningún archivo. Deteniendo el auto-guardado y cerrando la vista.")
            self.stop_auto_save()  # Detener el autoguardado
        else:
            print("Hay un archivo seleccionado. Realizando el guardado antes de cerrar.")
            self.auto_save()  # Realizar el autoguardado antes de cerrar
            self.stop_auto_save()  # Detener el autoguardado