import os
import re
import openpyxl
import pandas as pd

from components.get_analisis import obtener_datos_analisis
from components.get_path_resources import get_path_resources
from components.show_messages import show_error

class RangosModel:
    def __init__(self):
        self.headers1 = []
        self.all_data1 = []
        self.headers = []
        self.df = []

        self.all_data = []
        self.analysis_columns: dict[str, str] = obtener_datos_analisis()


    def predeterminado(self):
        ruta_archivo = get_path_resources("Rangos.xlsx")


        try:
            # Cargar el archivo Excel en un DataFrame
            df = pd.read_excel(ruta_archivo)

            # Actualizar headers y all_data
            self.headers = list(df.columns)
            self.all_data = df.fillna(" ").values.tolist()
            self.df = df
            
            return self.headers, self.all_data

        except Exception as e:
            show_error("Error", e)
            return None, None
        


    def obtener_datos(self):
        ruta_libro2 = get_path_resources("Libro2.xlsx")
        ruta_rangos = get_path_resources("Rangos.xlsx")

        try:
            # Cargar los datos actuales de Rangos.xlsx si existe
            if os.path.exists(ruta_rangos):
                df_rangos = pd.read_excel(ruta_rangos)
            else:
                df_rangos = pd.DataFrame(columns=["LOCALIDAD", "MUESTRA", "ANALISIS", "MINIMO", "MAXIMO"])

            # Cargar el nuevo archivo Libro2.xlsx
            df_nuevo = pd.read_excel(ruta_libro2)

            # Verificar columnas requeridas
            if "LOCALIDAD" not in df_nuevo.columns or "MUESTRA" not in df_nuevo.columns:
                raise ValueError("El archivo debe contener las columnas 'LOCALIDAD' y 'MUESTRA'.")

            # Crear la columna "ANALISIS" a partir de las columnas de análisis
            analysis_cols = [col for col in self.analysis_columns.values() if col in df_nuevo.columns]
            if analysis_cols:
                df_nuevo["ANALISIS"] = df_nuevo[analysis_cols].apply(
                    lambda row: ", ".join(row.dropna().astype(str)), axis=1
                )
                df_nuevo = df_nuevo.drop(columns=analysis_cols)

                # Expandir filas para cada análisis
                expanded_rows = []
                for _, row in df_nuevo.iterrows():
                    analisis = str(row["ANALISIS"]).split(",")
                    analisis = [a.strip() for a in analisis if a.strip()]  # Limpiar espacios

                    for analisis_item in analisis:
                        new_row = row.copy()
                        new_row["ANALISIS"] = analisis_item
                        expanded_rows.append(new_row)

                df_nuevo = pd.DataFrame(expanded_rows)

            # Añadir columnas "MINIMO" y "MAXIMO" con valores vacíos si no existen
            if "MINIMO" not in df_nuevo.columns:
                df_nuevo["MINIMO"] = ""
            if "MAXIMO" not in df_nuevo.columns:
                df_nuevo["MAXIMO"] = ""

            # Seleccionar solo las columnas relevantes
            df_nuevo = df_nuevo[["LOCALIDAD", "MUESTRA", "ANALISIS", "MINIMO", "MAXIMO"]]

            # Eliminar los registros que ya no existen en Libro2.xlsx
            df_rangos_filtrado = df_rangos.merge(df_nuevo[["LOCALIDAD", "MUESTRA", "ANALISIS"]], 
                                                on=["LOCALIDAD", "MUESTRA", "ANALISIS"], 
                                                how="inner")

            # Unir datos sin duplicados, manteniendo valores de MINIMO y MAXIMO existentes
            df_combinado = pd.concat([df_rangos_filtrado, df_nuevo]).drop_duplicates(
                subset=["LOCALIDAD", "MUESTRA", "ANALISIS"], keep="first"
            )

            # Guardar el archivo actualizado en Rangos.xlsx
            df_combinado.to_excel(ruta_rangos, index=False)

            # Retornar encabezados y datos actualizados
            self.headers1 = list(df_combinado.columns)
            self.all_data1 = df_combinado.values.tolist()
            return self.headers1, self.all_data1

        except Exception as e:
            raise show_error("Error", e)
            


    def quitar_unidades(self, texto):
        """Elimina unidades de medida de los nombres de análisis."""
        return re.sub(r"\s*\(.*?\)", "", texto).strip()

    def obtener_directorios(self):
        """Lee DirectoriosLocalidades.txt y obtiene los directorios por localidad."""
        ruta_txt = get_path_resources("DirectoriosLocalidades.txt")
        directorios = {}
        with open(ruta_txt, "r", encoding="utf-8") as file:
            for line in file:
                partes = line.strip().split(":", 1)
                if len(partes) == 2:
                    localidad, ruta = partes
                    print(partes)
                    directorios[localidad.strip().upper()] = ruta.strip()

        return directorios

    def buscar_en_excel(self, ruta, muestra, analisis_referencia):
        """Busca la muestra en la segunda fila y los análisis a partir de ahí, eliminando unidades de medida."""
        try:
            wb = openpyxl.load_workbook(ruta)
            sheet = wb.active

            # Buscar la muestra en la segunda fila
            muestra_col = None
            for col in range(1, sheet.max_column + 1):
                celda_valor = sheet.cell(row=2, column=col).value
                if celda_valor and str(celda_valor).strip().upper() == muestra:
                    muestra_col = col
                    break
            
            if not muestra_col:
                return None  # Muestra no encontrada

            # Buscar los análisis desde la columna de la muestra en adelante
            posiciones = {}
            for col in range(muestra_col, sheet.max_column + 1):
                celda_valor = sheet.cell(row=3, column=col).value  # Buscar en la fila siguiente
                if celda_valor:
                    analisis_limpio = self.quitar_unidades(str(celda_valor).strip().upper())  # Normalizar nombre de análisis
                    if analisis_limpio == analisis_referencia:
                        posiciones[analisis_limpio] = openpyxl.utils.get_column_letter(col)

            return posiciones  # Retorna las ubicaciones de los análisis

        except Exception as e:
            print(f"Error al leer {ruta}: {e}")
            return None

    def ubicaciones(self):
        print("hola3")
        ruta_archivo = get_path_resources("Rangos.xlsx")
        try:
            # Leer Rangos.xlsx
            print("Hola4")
            df = pd.read_excel(ruta_archivo)
            self.headers = list(df.columns)
            self.all_data = df.values.tolist()

            # Obtener los directorios por localidad
            directorios = self.obtener_directorios()
            print(directorios)

            # Guardar resultados de ubicación
            resultados = []

            for row in self.all_data:
                localidad = row[0].strip().upper()
                muestra = row[1].strip().upper()
                analisis = row[2].strip().upper()  # NO se limpia aquí

                if localidad in directorios:
                    ruta_excel = directorios[localidad]

                    # Buscar ubicaciones en el archivo Excel
                    posiciones = self.buscar_en_excel(ruta_excel, muestra, analisis)

                    if posiciones:
                        for analisis_encontrado, columna in posiciones.items():
                            resultados.append([localidad, muestra, analisis_encontrado, ruta_excel, columna])
            print(resultados)
            return resultados

        except Exception as e:
            print("Error:", str(e))
            return None
