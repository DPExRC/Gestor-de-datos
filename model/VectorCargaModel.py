import pandas as pd
from openpyxl import load_workbook

from components.get_analisis import obtener_datos_analisis

class VectorCargaModel:
    def __init__(self):
        self.headers = []
        self.all_data = []
        self.analysis_columns: dict[str, str] = obtener_datos_analisis()

        
        
    def load_file(self, file_path):
        """Cargar datos del archivo Excel seleccionado y psrocesarlos"""
        df = pd.read_excel(file_path)
        
    

        #df.rename(columns={"PLANTA": "LOCALIDAD"}, inplace=True)
        if any(col in df.columns for col in self.analysis_columns.values()):
            df["ANALISIS"] = df[list(self.analysis_columns.values())] \
                .apply(lambda row: ", ".join(row.dropna().astype(str)), axis=1)

            df = df.drop(columns=[col for col in self.analysis_columns.values() if col in df.columns])
            
            if "UBICACION" in df.columns:
                df["UBICACION"] = df["UBICACION"].fillna('')  # O usa dropna si quieres eliminar filas


        columns_order = [col for col in df.columns if col != "ANALISIS" and "UBICACION"] + ["ANALISIS"] 

        df = df[columns_order]


        self.headers = list(df.columns)
        
        self.all_data = df.values.tolist()

        return self.headers, self.all_data
    

    def export_to_excel(self, data, headers, file_path):
        """Exportar los datos a un archivo Excel"""
        df = pd.DataFrame(data, columns=headers)

        if "ANALISIS" in df.columns:
            for key, column in self.analysis_columns.items():
                df[column] = df["ANALISIS"].apply(
                    lambda x: next(
                        (val.strip() for val in str(x).split(",") if val.strip() == key), None
                    )
                )
            df = df.drop(columns=["ANALISIS"])

        df.to_excel(file_path, index=False)

        wb = load_workbook(file_path)
        ws = wb.active

        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col if cell.value)
            column_letter = col[0].column_letter
            ws.column_dimensions[column_letter].width = max_length + 2

        wb.save(file_path)

    def export_to_excel2(self, data, headers, file_path):
        """Exportar los datos a un archivo Excel"""
        df = pd.DataFrame(data, columns=headers)


        df.to_excel(file_path, index=False)

        wb = load_workbook(file_path)
        ws = wb.active

        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col if cell.value)
            column_letter = col[0].column_letter
            ws.column_dimensions[column_letter].width = max_length + 2

        wb.save(file_path)



